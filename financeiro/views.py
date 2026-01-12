from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from datetime import date, timedelta
import json
from .models import Grupo, ContaPagar
from .forms import GrupoForm, ContaPagarForm

# --- GRUPOS ---

class GrupoListView(LoginRequiredMixin, ListView):
    model = Grupo
    template_name = 'financeiro/grupo_list.html'
    context_object_name = 'grupos'

    def get_queryset(self):
        """Retorna apenas grupos do usuário logado."""
        return Grupo.objects.filter(usuario=self.request.user)

class GrupoCreateView(LoginRequiredMixin, CreateView):
    model = Grupo
    form_class = GrupoForm
    template_name = 'financeiro/grupo_form.html'
    success_url = reverse_lazy('grupo-list')

    def form_valid(self, form):
        """Associa automaticamente o grupo ao usuário logado."""
        form.instance.usuario = self.request.user
        return super().form_valid(form)

class GrupoUpdateView(LoginRequiredMixin, UpdateView):
    model = Grupo
    form_class = GrupoForm
    template_name = 'financeiro/grupo_form.html'
    success_url = reverse_lazy('grupo-list')

    def get_queryset(self):
        """Limita edição apenas aos grupos do usuário logado."""
        return Grupo.objects.filter(usuario=self.request.user)

class GrupoDeleteView(LoginRequiredMixin, DeleteView):
    model = Grupo
    template_name = 'financeiro/grupo_confirm_delete.html'
    success_url = reverse_lazy('grupo-list')

    def get_queryset(self):
        """Limita exclusão apenas aos grupos do usuário logado."""
        return Grupo.objects.filter(usuario=self.request.user)

class GrupoDetailView(LoginRequiredMixin, DetailView):
    model = Grupo
    template_name = 'financeiro/grupo_detail.html'
    context_object_name = 'grupo'

    def get_queryset(self):
        """Limita visualização apenas aos grupos do usuário logado."""
        return Grupo.objects.filter(usuario=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtro de Mês e Ano
        hoje = date.today()
        mes = int(self.request.GET.get('mes', hoje.month))
        ano = int(self.request.GET.get('ano', hoje.year))

        # Data usada para navegação
        data_atual = date(ano, mes, 1)
        
        # Calcular mês anterior
        if mes == 1:
            mes_anterior = 12
            ano_anterior = ano - 1
        else:
            mes_anterior = mes - 1
            ano_anterior = ano
            
        # Calcular próximo mês
        if mes == 12:
            mes_proximo = 1
            ano_proximo = ano + 1
        else:
            mes_proximo = mes + 1
            ano_proximo = ano

        # Filtrar contas do grupo para o mês/ano selecionado
        contas = self.object.contas.filter(
            data_vencimento__month=mes,
            data_vencimento__year=ano
        ).order_by('data_vencimento')

        # Totais
        total_previsto = contas.aggregate(Sum('valor'))['valor__sum'] or 0
        total_pago = contas.filter(pago=True).aggregate(Sum('valor'))['valor__sum'] or 0
        total_pendente = total_previsto - total_pago

        # Dados para gráfico de histórico (últimos 6 meses)
        historico_labels = []
        historico_previsto = []
        historico_pago = []
        
        MESES_PT = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                    'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        
        for i in range(5, -1, -1):  # 5 a 0 (6 meses, do mais antigo ao atual)
            # Calcular mês/ano para cada iteração
            m = mes - i
            a = ano
            while m <= 0:
                m += 12
                a -= 1
            
            historico_labels.append(f"{MESES_PT[m]}/{a}")
            
            # Buscar totais deste mês
            contas_mes = self.object.contas.filter(
                data_vencimento__month=m,
                data_vencimento__year=a
            )
            prev = contas_mes.aggregate(Sum('valor'))['valor__sum'] or 0
            pag = contas_mes.filter(pago=True).aggregate(Sum('valor'))['valor__sum'] or 0
            
            historico_previsto.append(float(prev))
            historico_pago.append(float(pag))

        context.update({
            'contas': contas,
            'mes_atual': mes,
            'ano_atual': ano,
            'data_atual': data_atual,
            'mes_anterior': mes_anterior,
            'ano_anterior': ano_anterior,
            'mes_proximo': mes_proximo,
            'ano_proximo': ano_proximo,
            'total_previsto': total_previsto,
            'total_pago': total_pago,
            'total_pendente': total_pendente,
            'today': date.today(),
            # Dados para gráficos (JSON)
            'chart_historico_labels': json.dumps(historico_labels),
            'chart_historico_previsto': json.dumps(historico_previsto),
            'chart_historico_pago': json.dumps(historico_pago),
        })
        return context

# --- CONTAS A PAGAR ---

class ContaPagarCreateView(LoginRequiredMixin, CreateView):
    model = ContaPagar
    form_class = ContaPagarForm
    template_name = 'financeiro/contapagar_form.html'
    
    def get_initial(self):
        initial = super().get_initial()
        # Preenche o grupo se passado na URL (verifica se pertence ao usuário)
        grupo_id = self.kwargs.get('grupo_id')
        if grupo_id:
            initial['grupo'] = get_object_or_404(Grupo, pk=grupo_id, usuario=self.request.user)
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grupo_id = self.kwargs.get('grupo_id')
        if grupo_id:
            context['grupo'] = get_object_or_404(Grupo, pk=grupo_id, usuario=self.request.user)
        return context

    def get_success_url(self):
        return reverse('grupo-detail', kwargs={'pk': self.object.grupo.pk})

class ContaPagarUpdateView(LoginRequiredMixin, UpdateView):
    model = ContaPagar
    form_class = ContaPagarForm
    template_name = 'financeiro/contapagar_form.html'

    def get_queryset(self):
        """Limita edição às contas de grupos do usuário logado."""
        return ContaPagar.objects.filter(grupo__usuario=self.request.user)

    def get_success_url(self):
        return reverse('grupo-detail', kwargs={'pk': self.object.grupo.pk})

class ContaPagarDeleteView(LoginRequiredMixin, DeleteView):
    model = ContaPagar
    template_name = 'financeiro/contapagar_confirm_delete.html'

    def get_queryset(self):
        """Limita exclusão às contas de grupos do usuário logado."""
        return ContaPagar.objects.filter(grupo__usuario=self.request.user)

    def get_success_url(self):
        return reverse('grupo-detail', kwargs={'pk': self.object.grupo.pk})


# --- EXPORTAÇÃO PDF / EXCEL ---

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from io import BytesIO

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MESES_PT = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']


@login_required
def exportar_pdf(request, pk):
    """Exporta o resumo mensal de um grupo em formato PDF."""
    grupo = get_object_or_404(Grupo, pk=pk, usuario=request.user)
    
    hoje = date.today()
    mes = int(request.GET.get('mes', hoje.month))
    ano = int(request.GET.get('ano', hoje.year))
    
    contas = ContaPagar.objects.filter(
        grupo=grupo,
        data_vencimento__month=mes,
        data_vencimento__year=ano
    ).order_by('data_vencimento')
    
    # Totais
    total_previsto = contas.aggregate(Sum('valor'))['valor__sum'] or 0
    total_pago = contas.filter(pago=True).aggregate(Sum('valor'))['valor__sum'] or 0
    total_pendente = total_previsto - total_pago
    
    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           leftMargin=2*cm, rightMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=6
    )
    elements.append(Paragraph(f"{grupo.nome}", title_style))
    
    # Subtítulo (mês/ano)
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.grey,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Resumo de {MESES_PT[mes]} de {ano}", subtitle_style))
    
    # Resumo financeiro
    resumo_data = [
        ['Total Previsto', 'Total Pago', 'Pendente'],
        [f'R$ {total_previsto:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
         f'R$ {total_pago:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
         f'R$ {total_pendente:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')]
    ]
    
    resumo_table = Table(resumo_data, colWidths=[5.5*cm, 5.5*cm, 5.5*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#0d6efd')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#198754')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#ffc107')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
        ('TEXTCOLOR', (2, 0), (2, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(resumo_table)
    elements.append(Spacer(1, 20))
    
    # Tabela de contas
    if contas.exists():
        elements.append(Paragraph("Detalhamento das Contas", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        table_data = [['Status', 'Vencimento', 'Descrição', 'Valor']]
        
        for conta in contas:
            status = '✓ Pago' if conta.pago else '○ Pendente'
            vencimento = conta.data_vencimento.strftime('%d/%m/%Y')
            valor = f'R$ {conta.valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            table_data.append([status, vencimento, conta.descricao, valor])
        
        contas_table = Table(table_data, colWidths=[2.5*cm, 3*cm, 8*cm, 3*cm])
        contas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(contas_table)
    else:
        elements.append(Paragraph("Nenhuma conta cadastrada para este mês.", styles['Normal']))
    
    doc.build(elements)
    
    # Response
    buffer.seek(0)
    filename = f"resumo_{grupo.nome.lower().replace(' ', '_')}_{MESES_PT[mes].lower()}_{ano}.pdf"
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def exportar_excel(request, pk):
    """Exporta o resumo mensal de um grupo em formato Excel."""
    grupo = get_object_or_404(Grupo, pk=pk, usuario=request.user)
    
    hoje = date.today()
    mes = int(request.GET.get('mes', hoje.month))
    ano = int(request.GET.get('ano', hoje.year))
    
    contas = ContaPagar.objects.filter(
        grupo=grupo,
        data_vencimento__month=mes,
        data_vencimento__year=ano
    ).order_by('data_vencimento')
    
    # Totais
    total_previsto = contas.aggregate(Sum('valor'))['valor__sum'] or 0
    total_pago = contas.filter(pago=True).aggregate(Sum('valor'))['valor__sum'] or 0
    total_pendente = total_previsto - total_pago
    
    # Criar Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESES_PT[mes]} {ano}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    success_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    warning_fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
    table_header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = f"{grupo.nome} - Resumo de {MESES_PT[mes]} de {ano}"
    ws['A1'].font = Font(bold=True, size=16, color="0d6efd")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Resumo 
    ws['A3'] = "Total Previsto"
    ws['B3'] = "Total Pago"
    ws['C3'] = "Pendente"
    
    for col, fill in [('A', header_fill), ('B', success_fill), ('C', warning_fill)]:
        ws[f'{col}3'].fill = fill
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF" if col != 'C' else "000000")
        ws[f'{col}3'].alignment = Alignment(horizontal='center')
        ws[f'{col}3'].border = thin_border
    
    ws['A4'] = total_previsto
    ws['B4'] = total_pago
    ws['C4'] = total_pendente
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}4'].number_format = 'R$ #,##0.00'
        ws[f'{col}4'].font = Font(bold=True, size=12)
        ws[f'{col}4'].alignment = Alignment(horizontal='center')
        ws[f'{col}4'].border = thin_border
    
    # Tabela de contas
    ws['A6'] = "Status"
    ws['B6'] = "Vencimento"
    ws['C6'] = "Descrição"
    ws['D6'] = "Valor"
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}6'].fill = table_header_fill
        ws[f'{col}6'].font = header_font
        ws[f'{col}6'].alignment = Alignment(horizontal='center')
        ws[f'{col}6'].border = thin_border
    
    row = 7
    for conta in contas:
        ws[f'A{row}'] = "✓ Pago" if conta.pago else "○ Pendente"
        ws[f'B{row}'] = conta.data_vencimento.strftime('%d/%m/%Y')
        ws[f'C{row}'] = conta.descricao
        ws[f'D{row}'] = conta.valor
        ws[f'D{row}'].number_format = 'R$ #,##0.00'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center' if col in ['A', 'B'] else 'left')
        
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    
    # Response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"resumo_{grupo.nome.lower().replace(' ', '_')}_{MESES_PT[mes].lower()}_{ano}.xlsx"
    
    response = HttpResponse(
        buffer, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
